"""
name: xlsx_split.py
description: Split excel sheet into multiple sheets based on column value.
author: hdaojin
version: 1.0.0
date: 2022.10.20
update: 2022.10.20
usage: python3 xlsx_split.py --help
"""

from argparse import ArgumentParser, MetavarTypeHelpFormatter
from pathlib import Path
from shutil import copyfile
from openpyxl import load_workbook, Workbook

Version = '1.0.0'

def split_sheet(input_file, sheet_name, start_row=1, start_col=1, end_row=None, end_col=None):
    # Load workbook
    wb = load_workbook(input_file)
    # Get sheet, if sheet_name is not specified, get the first sheet.
    if not sheet_name:
        ws = wb.active

    # Get max row and column
    end_row = end_row or ws.max_row
    end_col = end_col or ws.max_column

    # Get data
    data = []
    for raw in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        i = [cell.value for cell in raw]
        data.append(i)

    # Split data
    # Use original file as template
    wb_split = wb
    ws_template = wb_split.active

    # Write data to new worksheet that was split by column value from the original worksheet

    for i in data:
        ws_new = wb_split.copy_worksheet(ws_template)
        ws_new.title = i[0]
        # Delete data from template worksheet if it is not header.
        ws_new.delete_rows(start_row, end_row)
        ws_new.append(i)
    # new workbook's name
    new_wb_name = Path(input_file).stem + '_split.xlsx'
    new_wb_file = Path(input_file).parent / new_wb_name
    wb_split.save(new_wb_file)
    print(f'Finished, saved to {new_wb_file} !')

if __name__ == '__main__':
    # parser = ArgumentParser(description='Split excel sheet into multiple sheets based on column value.', formatter_class=RawTextHelpFormatter)
    parser = ArgumentParser(description='Split excel sheet into multiple sheets based on column value.', formatter_class=MetavarTypeHelpFormatter)
    parser.add_argument('-v', '--version', action='version', version=f'%(prog)s {Version}')
    parser.add_argument('-s', '--sheet', metavar='SHEET', type=str, help='Sheet name, default is active sheet.')
    parser.add_argument('-r', '--start-row', type=int, default=1, help='Start row, default is 1')
    parser.add_argument('-c', '--start-col', type=int, default=1, help='Start column, default is 1')
    parser.add_argument('-e', '--end-row', type=int, help='End row, default is max row')
    parser.add_argument('-d', '--end-col', type=int, help='End column, default is max column')
    parser.add_argument('file', metavar='FILE', type=Path, help='Input excel file')
    args = parser.parse_args()

    split_sheet(args.file, args.sheet, args.start_row, args.start_col, args.end_row, args.end_col)